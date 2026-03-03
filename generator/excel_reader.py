"""Parse Excel data file into a flat dictionary for placeholder replacement."""

import re
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

from .constants import MAX_SERVICES, MAX_CITIES, MAX_UNITS


def read_excel(filepath: str) -> Dict[str, str]:
    """Read the Excel data file and return a flat dict of placeholder -> value.

    Reads the Quick Copy sheet (primary data source) and the Tracking & Schema
    sheet for GTM/analytics IDs. Also reads Company Data, Images, City 1-5,
    and Size Guide sheets to capture any data not in Quick Copy.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    data = {}

    # Read Quick Copy sheet (primary data source)
    if "Quick Copy" in wb.sheetnames:
        _read_kv_sheet(wb["Quick Copy"], data)

    # Read supplementary sheets (won't overwrite Quick Copy values)
    for sheet_name in ["Company Data", "Images", "Tracking & Schema", "Size Guide"]:
        if sheet_name in wb.sheetnames:
            _read_kv_sheet(wb[sheet_name], data, overwrite=False)

    # Read City 1-5 sheets
    for n in range(1, 6):
        sheet_name = f"City {n}"
        if sheet_name in wb.sheetnames:
            _read_kv_sheet(wb[sheet_name], data, overwrite=False)

    wb.close()

    # Detect counts
    data["_service_count"] = str(_detect_service_count(data))
    data["_city_count"] = str(_detect_city_count(data))
    data["_enabled_units"] = ",".join(str(u) for u in _detect_enabled_units(data))

    # Ensure CURRENT_YEAR is set
    if "CURRENT_YEAR" not in data:
        from datetime import datetime
        data["CURRENT_YEAR"] = str(datetime.now().year)

    # Ensure DOMAIN ends with /
    if "DOMAIN" in data and not data["DOMAIN"].endswith("/"):
        data["DOMAIN"] = data["DOMAIN"] + "/"

    # Ensure DOMAIN starts with https://
    if "DOMAIN" in data and not data["DOMAIN"].startswith("http"):
        data["DOMAIN"] = "https://" + data["DOMAIN"]

    # Generate default content for template placeholders missing from Excel
    _fill_content_defaults(data)

    return data


def _read_kv_sheet(ws, data: Dict[str, str], overwrite: bool = True):
    """Read a key-value sheet (col A = placeholder, col B = value) into data dict."""
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is None:
            continue
        key = row[0]
        value = row[1]
        clean_key = _strip_braces(str(key).strip())
        if not clean_key:
            continue
        # Include None values as empty string so placeholder gets replaced
        clean_value = _clean_value(value) if value is not None else ""
        if overwrite or clean_key not in data:
            data[clean_key] = clean_value


def _fill_content_defaults(data: Dict[str, str]):
    """Fill in default content for template placeholders not provided in Excel.

    These are content-level fields (meta descriptions, FAQ Q&A, service
    descriptions/features, city descriptions) that the templates expect
    but the Excel Quick Copy sheet doesn't always include.
    """
    biz = data.get("BUSINESS_NAME", "Our Storage Facility")
    city = data.get("PRIMARY_CITY", "our area")
    state = data.get("STATE_CODE", "")
    tagline = data.get("TAGLINE", "")
    location = f"{city}, {state}" if state else city

    # --- Page meta descriptions ---
    if not data.get("ABOUT_META_DESC"):
        desc = f"Learn about {biz} in {location}."
        if tagline:
            desc += f" {tagline}"
        data["ABOUT_META_DESC"] = desc

    if not data.get("FAQS_META_DESC"):
        data["FAQS_META_DESC"] = (
            f"Frequently asked questions about self storage at {biz} in {location}. "
            f"Find answers about unit sizes, access hours, security, and more."
        )

    if not data.get("SERVICES_META_DESC"):
        data["SERVICES_META_DESC"] = (
            f"Explore storage services at {biz} in {location}. "
            f"From self-storage to climate-controlled units, we have solutions for every need."
        )

    if not data.get("SIZEGUIDE_META_DESC"):
        data["SIZEGUIDE_META_DESC"] = (
            f"Find the right storage unit size at {biz}. "
            f"View our size guide with dimensions and recommendations for your belongings."
        )

    # --- Service descriptions and features ---
    service_count = _detect_service_count(data)
    for n in range(1, service_count + 1):
        svc_name = data.get(f"SERVICE_{n}_NAME", f"Service {n}")

        if not data.get(f"SERVICE_{n}_META_DESC"):
            data[f"SERVICE_{n}_META_DESC"] = (
                f"{svc_name} at {biz} in {location}. "
                f"Secure, accessible, and affordable storage solutions."
            )

        if not data.get(f"SERVICE_{n}_SHORT_DESC"):
            data[f"SERVICE_{n}_SHORT_DESC"] = (
                f"Professional {svc_name.lower()} solutions in {location}."
            )

        if not data.get(f"SERVICE_{n}_FULL_DESC"):
            data[f"SERVICE_{n}_FULL_DESC"] = (
                f"{biz} offers {svc_name.lower()} to meet your storage needs in {location}. "
                f"Our facility features modern security systems, convenient access hours, "
                f"and well-maintained units to keep your belongings safe and accessible."
            )

        # Default features based on common self-storage offerings
        default_features = [
            "24/7 Video Surveillance",
            "Convenient Access Hours",
            "Well-Lit Facility",
            "Month-to-Month Leases",
        ]
        for i in range(1, 5):
            if not data.get(f"SERVICE_{n}_FEATURE_{i}"):
                data[f"SERVICE_{n}_FEATURE_{i}"] = default_features[i - 1]

    # --- City descriptions ---
    city_count = _detect_city_count(data)
    address = data.get("STREET_ADDRESS", "")
    for n in range(1, city_count + 1):
        city_name = data.get(f"CITY_{n}_NAME", f"City {n}")
        if not data.get(f"CITY_{n}_FULL_DESC"):
            parts = [
                f"{biz} is proud to serve residents of {city_name}, {state}.",
            ]
            if address and city:
                parts.append(
                    f"Our facility at {address} in {city} offers easy access "
                    f"and a range of storage options for {city_name} residents."
                )
            parts.append(
                f"Whether you need short-term storage during a move or long-term "
                f"space for your belongings, we have the right unit for you."
            )
            data[f"CITY_{n}_FULL_DESC"] = " ".join(parts)

    # --- FAQ Q&A (8 generic self-storage FAQs) ---
    _default_faqs = [
        (
            "What size storage unit do I need?",
            f"The right unit size depends on what you're storing. A 5x5 unit works for boxes "
            f"and small furniture, while a 10x20 can hold the contents of a multi-bedroom home. "
            f"Visit our size guide or contact us at {data.get('PHONE_DISPLAY', 'our office')} "
            f"for a personalized recommendation."
        ),
        (
            "What are your access hours?",
            f"Our facility offers convenient access hours to fit your schedule. "
            f"Contact us at {data.get('PHONE_DISPLAY', 'our office')} for current hours "
            f"and gate access details."
        ),
        (
            "Is my stuff safe in your storage units?",
            f"Absolutely. {biz} features 24/7 video surveillance, secure gated access, "
            f"and individually alarmed units to ensure your belongings are protected around the clock."
        ),
        (
            "Do you offer climate-controlled storage?",
            f"Yes, we offer climate-controlled units that maintain consistent temperature and humidity "
            f"levels, ideal for furniture, electronics, documents, and other sensitive items."
        ),
        (
            "How do I rent a storage unit?",
            f"Renting is easy! You can reserve a unit online through our website, call us at "
            f"{data.get('PHONE_DISPLAY', 'our office')}, or visit our facility in person. "
            f"We offer flexible month-to-month leases with no long-term commitment required."
        ),
        (
            "Can I access my unit anytime?",
            f"We provide generous access hours so you can reach your belongings when you need them. "
            f"Check our access schedule or call for details about after-hours access options."
        ),
        (
            "Do you have vehicle or RV storage?",
            f"Yes, {biz} offers parking and storage options for cars, boats, RVs, and other vehicles. "
            f"Contact us to learn about available spaces and sizes."
        ),
        (
            "What payment methods do you accept?",
            f"We accept major credit cards, debit cards, and electronic payments for your convenience. "
            f"Auto-pay is also available so you never miss a payment."
        ),
    ]
    for i, (question, answer) in enumerate(_default_faqs, start=1):
        if not data.get(f"FAQ_{i}_QUESTION"):
            data[f"FAQ_{i}_QUESTION"] = question
        if not data.get(f"FAQ_{i}_ANSWER"):
            data[f"FAQ_{i}_ANSWER"] = answer


def _strip_braces(key: str) -> str:
    """Remove {{ }} from placeholder key."""
    if key.startswith("{{") and key.endswith("}}"):
        return key[2:-2].strip()
    return key


def _clean_value(value) -> str:
    """Clean cell value: strip .0 from numbers, convert to string."""
    if isinstance(value, float):
        # Strip .0 from integers stored as floats (e.g., ZIP codes)
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _detect_service_count(data: Dict[str, str]) -> int:
    """Count how many services are defined (1-6)."""
    count = 0
    for n in range(1, MAX_SERVICES + 1):
        if data.get(f"SERVICE_{n}_NAME"):
            count = n
        else:
            break
    return count


def _detect_city_count(data: Dict[str, str]) -> int:
    """Count how many cities are defined (1-5)."""
    count = 0
    for n in range(1, MAX_CITIES + 1):
        if data.get(f"CITY_{n}_NAME"):
            count = n
        else:
            break
    return count


def _detect_enabled_units(data: Dict[str, str]) -> List[int]:
    """Return list of unit numbers that are enabled."""
    enabled = []
    for n in range(1, MAX_UNITS + 1):
        val = data.get(f"UNIT_{n}_ENABLED", "").lower()
        if val in ("yes", "true", "1", "y"):
            enabled.append(n)
    return enabled


def get_service_count(data: Dict[str, str]) -> int:
    """Get the number of services from data dict."""
    return int(data.get("_service_count", "0"))


def get_city_count(data: Dict[str, str]) -> int:
    """Get the number of cities from data dict."""
    return int(data.get("_city_count", "0"))


def get_enabled_units(data: Dict[str, str]) -> List[int]:
    """Get list of enabled unit numbers from data dict."""
    units_str = data.get("_enabled_units", "")
    if not units_str:
        return []
    return [int(u) for u in units_str.split(",")]
